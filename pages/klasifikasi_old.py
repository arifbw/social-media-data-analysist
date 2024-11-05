import streamlit as st
import pandas as pd
import plotly.express as px
from PIL import Image
import re
import requests
from openpyxl import load_workbook
# from stqdm import stqdm
from io import BytesIO
from datetime import datetime
from streamlit_condition_tree import condition_tree, config_from_dataframe
import time
import pymongo
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader
import page_service as ps
from streamlit_tags import st_tags

# init page
st.set_page_config(
    page_title="Social Media Data Analysist",
    page_icon="📊",
    layout="wide"
)

img_logo = Image.open('logo_pasker.png')
st.logo(img_logo, size="large")

with open('config.yaml') as file:
    config_yaml = yaml.load(file, Loader=SafeLoader)

# password = stauth.Hasher.hash_passwords(config_yaml['credentials'])
# st.write(password)

authenticator = stauth.Authenticate(
    config_yaml['credentials'],
    config_yaml['cookie']['name'],
    config_yaml['cookie']['key'],
    config_yaml['cookie']['expiry_days'],
    auto_hash=False
)

authenticator.login(location="unrendered")

st.session_state["auth_obj"] = authenticator

client = None
db = None

@st.cache_resource
def init_connection():
    return pymongo.MongoClient(**st.secrets["mongo"])

@st.cache_data
def kalkulasi_banyak_row(uploaded_file):
    df = pd.read_excel(uploaded_file, sheet_name="Media Sosial")
    return len(df)

def classify_job_category(caption, categories):
        contains_others = bool(re.search("Lainnya", caption, re.IGNORECASE))

        found_categories = []
        for category in categories:
            if re.search(category, caption, re.IGNORECASE):
                found_categories.append(category)

        if found_categories:
            return found_categories[0]

        return 'Lainnya'

def classify_job_category2(caption, categories):
    matched_categories = []

    # Iterate through each category
    for category in categories["category_data"]:
        # Check if any threshold word is present in the caption (ignore case)
        for keyword in category["threshold"]:
            if re.search(r"\b" + re.escape(keyword) + r"\b", caption, re.IGNORECASE):
            # if re.search(r"(?<!\w)" + re.escape(keyword) + r"(?!\w)", caption, re.IGNORECASE):
            # if re.search(r"(?<![a-zA-Z])" + re.escape(keyword) + r"(?![a-zA-Z])", caption, re.IGNORECASE):
                matched_categories.append(category["name"])
                break  # Move to the next category once a match is found

    if not matched_categories:
        matched_categories.append("tdk_ada_informasi")

    if "tdk_ada_informasi" in matched_categories and len(matched_categories) > 1:
        matched_categories.remove("tdk_ada_informasi")

    return matched_categories

def classify_akun(nama_akun, is_pers_value):
    kat_akun = "Akun Netizen"
    pattern = re.compile(r'|'.join(kat_akun_loker), re.IGNORECASE)

    if(is_pers_value):
        return "Akun Pers"
    
    if pattern.search(nama_akun):
        return "Akun Loker"
    
    return kat_akun

def kalkulasi_persentase_lowongan(params):
    result = 0
    params.insert(0, params.pop())

    for i, param in enumerate(params):
        if(i==0 and param=="Akun Loker"):
            result += nilai_bobot[0]
        elif(i != 0 and "tdk_ada_informasi" not in param):
            result += nilai_bobot[i]

    return result

def rand_persen_wait(persen):
    result = False
    
    if(persen > 5 and persen < 9):
        result = True
    elif(persen > 10 and persen < 15):
        result = True
    elif(persen > 20 and persen < 25):
        result = True
    elif(persen > 30 and persen < 35):
        result = True
    elif(persen > 40 and persen < 45):
        result = True
    elif(persen > 50 and persen < 55):
        result = True
    elif(persen > 60 and persen < 65):
        result = True
    elif(persen > 70 and persen < 75):
        result = True
    elif(persen > 80 and persen < 85):
        result = True
    elif(persen > 90 and persen < 95):
        result = True

    return result

def process_chunk(chunk, kd, collection_name, progress_bar):
    for index, row in chunk.iterrows():
        # Get the 'Konten' value once for this row
        konten_value = str(row['Konten'])
        akun_value = str(row['Akun/Judul'])
        is_pers_value = str(row['Jenis Akun']) == "Pers"
        data_params_kalkulasi_persentase_lowongan = []
        
        # Calculate and set classifications for each item in kd
        for item in kd:
            # Assuming item["kamus_data"] is the name of the new column
            column_name = item["kamus_data"]

            if column_name not in chunk.columns:
                # Create the column if it doesn't exist
                chunk[column_name] = None
            
            # Apply classify_job_category2 and assign result to the specific cell
            result_1 = classify_job_category2(konten_value, item)
            chunk.at[index, column_name] = result_1
            data_params_kalkulasi_persentase_lowongan.append(result_1)


        # Mengkliasifikasikan Tipe Akun
        if "Klasifikasi Akun" not in chunk.columns:
            chunk["Klasifikasi Akun"] = None

        result_2 = classify_akun(akun_value, is_pers_value)
        chunk.at[index, "Klasifikasi Akun"] = result_2
        data_params_kalkulasi_persentase_lowongan.append(result_2)


        # Menghitung Persentase Lowongan
        if "Persentase Lowongan" not in chunk.columns:
            chunk["Persentase Lowongan"] = None

        chunk.at[index, "Persentase Lowongan"] = kalkulasi_persentase_lowongan(data_params_kalkulasi_persentase_lowongan)


        progress_bar.progress(index/len(chunk), f"Menganalisa data baris ke-{index} dari {len(chunk)} data.")

    progress_bar.empty()

    collection = db[collection_name]
    data_dict = chunk.to_dict(orient="records")
    collection.insert_many(data_dict)

def clean_location_name(name):
    """
    Removes prefixes like 'Desa', 'Kecamatan', 'Kota', etc. and returns the cleaned name.
    """
    # Remove any prefix like Desa, Kota, Kecamatan, etc.
    prefixes = ['KOTA ', 'KAB. ', 'DAERAH ISTIMEWA ', 'KOTA ADM. ']
    for prefix in prefixes:
        if name.startswith(prefix):
            name = name.replace(prefix, "")
    return name.strip()

# Function to map location from Konten based on location data
def map_location(caption, location_df):
    """
    Finds the first matching location in the caption by comparing to the cleaned location names.
    """
    for location in location_df['cleaned_nama']:
        # Check if the location is found in the caption (ignore case)
        if re.search(location, caption, re.IGNORECASE):
            # If found, return the corresponding full location name
            return location_df[location_df['cleaned_nama'] == location]['nama'].values[0]
    # If no location is found, return 'Unknown'
    return 'Unknown'

job_categories = None
uploaded_file = None

keywords = None
kat_akun_loker = None
nilai_bobot = []

def categorize_job_post(caption):
    caption = caption.lower()  # Convert caption to lowercase for case-insensitive matching
    caption = re.sub(r'[^\w\s]', '', caption)  # Remove punctuation

    for kd in job_categories:
        # Iterate over categories and their keywords
        for category in kd['category_data']:
            name = category['name']

            for keyword in category['threshold']:
                if keyword in caption:
                    return name

    return 'uncategorized'

def get_kamus_data():
    kamus_data_xls = pd.ExcelFile('kamus_data.xlsx')
    result = []
    
    for sheet_name in keywords:
        # Load the sheet into a DataFrame
        df = pd.read_excel(kamus_data_xls, sheet_name=sheet_name)

        # Remove columns that are completely empty (all NaN values)
        df = df.dropna(axis=1, how='all')
        
        # Remove rows that are completely empty (all NaN values)
        df = df.dropna(axis=0, how='all')
        
        
        # Initialize a list to hold column data for this sheet
        column_data = []

        # Iterate through each column in the DataFrame
        for column in df.columns:
            # Get the column data as a list
            list_data = df[column].dropna().tolist()
            
            # Add column name and list of data to column_data list
            if list_data:
                column_data.append({
                    "name": column,
                    "threshold": list_data
                })
            
        # Append the sheet data to result
        if column_data:
            result.append({
                "kamus_data": sheet_name,
                "category_data": column_data
            })

    return result

def add_filter_component(df):
    with st.expander("Filter Data"):
        date_range = st.date_input("Select Tanggal Publikasi Range", [])
        topik_filter = st.multiselect("Select Topik", df['Topik'].unique())
        sentimen_filter = st.multiselect("Select Sentimen", df['Sentimen'].unique())
    
    # Filter by date (Tanggal Publikasi)
    if len(date_range) == 2:
        start_date, end_date = date_range
        df['Tanggal Publikasi'] = pd.to_datetime(df['Tanggal Publikasi'], errors='coerce')
        df = df[(df['Tanggal Publikasi'] >= pd.Timestamp(start_date)) & (df['Tanggal Publikasi'] <= pd.Timestamp(end_date))]

    # Filter by Topik
    if topik_filter:
        df = df[df['Topik'].isin(topik_filter)]

    # Filter by Sentimen
    if sentimen_filter:
        df = df[df['Sentimen'].isin(sentimen_filter)]

    return df

def setup_data_stats_and_sentiment(data, data2):
    st.sidebar.markdown(f"""
        <style>
            .custom-list li {{
                padding: 3px;
                font-size: 16px;
                border-bottom: 2px solid #cecece;
                list-style-type: none;
                margin-left: 0px;
            }}

            .title_sidebar{{
                color: #FF3333;
                font-weight: 900 !important;
            }}
                        
            iframe[title="streamlit_condition_tree.streamlit_condition_tree"]{{
                border: 3px #cccccc dashed;
                border-radius: 10px;
                padding: 20px 20px;
                box-sizing: content-box;
            }}
        </style>
        <h3 class="title_sidebar">Sosmed Data</h3>
        <ul class="custom-list">
            <li><i class="fa-brands fa-linkedin"></i> <strong>LinkedIn</strong>: {data['linkedin']:,} data</li>
            <li><i class="fa-brands fa-facebook"></i> <strong>Facebook</strong>: {data['facebook']:,} data</li>
            <li><i class="fa-brands fa-twitter"></i> <strong>Twitter</strong>: {data['twitter']:,} data</li>
            <li><i class="fa-brands fa-youtube"></i> <strong>YouTube</strong>: {data['youtube']:,} data</li>
            <li><i class="fa-brands fa-tiktok"></i> <strong>TikTok</strong>: {data['tiktok']:,} data</li>
            <li><i class="fa-brands fa-square-instagram"></i> <strong>Instagram</strong>: {data['instagram']:,} data</li>
        </ul>
    """, unsafe_allow_html=True)

    st.sidebar.markdown(f"""
        <h3 class="title_sidebar">Post Sentiment</h3>
        <ul class="custom-list">
            <li><i class="fa-solid fa-face-tired"></i> <strong>Negatif</strong>: {data2['negatif']:,} data</li>
            <li><i class="fa-solid fa-face-surprise"></i> <strong>Netral</strong>: {data2['netral']:,} data</li>
            <li><i class="fa-solid fa-face-smile"></i> <strong>Positif</strong>: {data2['positif']:,} data</li>
        </ul>
    """, unsafe_allow_html=True)

def generate_url(base_url, from_date, to_date, date_type="date", sort="date"):
    source = "ZmFjZWJvb2ssdHdpdHRlcixpbnN0YWdyYW0sdGlrdG9rLHlvdXR1YmUsZm9ydW0sYmxvZyxsaW5rZWRpbg=="
    now = datetime.now()
    cur_time = now.strftime("%Y%m%d%H%M%S")

    return f"{base_url}?from={from_date}%2000:00:00&to={to_date}%2023:59:59&date_type={date_type}&sort={sort}&sources={source}&time={cur_time}"

dynamic_url = None

@st.dialog("Download disini :")
def show_dynamic_url(url):
    st.write(f"{url}")

@st.dialog("Upload New Data :")
def update_kamus_data():
    data_baru = st.file_uploader("Upload Kamus Data :", type=["xlsx", "xls"])

    left, middle, right = st.columns(3)

    with right:
       if st.button("Lanjutkan", use_container_width=True, disabled=not data_baru):
           # Save the uploaded file to replace the existing one
            with open("kamus_data.xlsx", "wb") as f:
                f.write(data_baru.getbuffer())
                st.rerun()


def convert_df_to_excel(df):
    # Create a copy of the DataFrame to avoid modifying the original
    df_copy = df.copy()
    
    # Convert any column with list values to comma-separated strings
    for col in df_copy.columns:
        df_copy[col] = df_copy[col].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)
    
    # Write the DataFrame to an Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_copy.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)
    return output

def eksekusi_excel2(df, kd):

    # for item in stqdm(kd, desc="Mengalisa data semantik konten ...", backend=False, frontend=True):
        # df[item["kamus_data"]] = df['Konten'].progress_apply(lambda x: classify_job_category2(str(x), item))

    formatted_array = [item["kamus_data"] for item in kd]
    st.dataframe(df[['Konten', 'Url'] + formatted_array])

    
    col1, col2, col3, col4 = st.columns(4, gap="large")

    with col4:
        st.download_button(
            label="Download Excel file",
            use_container_width=True,
            type="primary",
            icon=":material/download:",
            data=convert_df_to_excel(df[['Konten', 'Url'] + formatted_array]),
            file_name='hasil_proses.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return
    
    col1, col2 = st.columns(2, gap="large")

    with col1:
        for item in formatted_array:
            category_count = df[item].explode().value_counts().reset_index()
            category_count.columns = [item, 'Count']
            category_count_filtered = category_count[category_count[item] != 'tdk_ada_informasi']

            # Pie chart
            fig_pie_category = px.pie(category_count_filtered, names=item, values='Count', title=f'{item} Distribution in Captions')
            st.plotly_chart(fig_pie_category)

    with col2:
        for item in formatted_array:
            category_count = df[item].explode().value_counts().reset_index()
            category_count.columns = [item, 'Count']
            category_count_filtered = category_count[category_count[item] != 'tdk_ada_informasi']
            
            # Horizontal bar 
            fig_bar_category = px.bar(category_count_filtered, x='Count', y=item, orientation='h', title=f'{item} Distribution (Horizontal Bar Chart)')
            st.plotly_chart(fig_bar_category)


def eksekusi_excel(tab1, tab2, df):
    with tab1:
        with st.spinner('Memproses data, mohon tunggu sebentar ...'):
            df = add_filter_component(df)
            st.dataframe(df)

            col1, col2 = st.columns(2, gap="large")

            with col1:
                # Bar chart of Sentimen distribution
                st.write("### Sentimen Distribution")
                sentimen_count = df['Sentimen'].value_counts().reset_index()
                sentimen_count.columns = ['Sentimen', 'Count']
                fig_sentimen = px.bar(sentimen_count, x='Sentimen', y='Count', color='Sentimen')
                st.plotly_chart(fig_sentimen)

                st.write("### Engagement by Akun/Judul")
                engagement_by_account = df.groupby('Akun/Judul').agg({'Engagement':'sum'}).reset_index().sort_values(by='Engagement', ascending=False)
                fig_account = px.bar(engagement_by_account, x='Akun/Judul', y='Engagement', title='Top Accounts by Engagement')
                st.plotly_chart(fig_account)
            
            with col2:
                # Heatmap for Sentiment and Emotion
                st.write("### Sentiment vs Emotion Heatmap")
                heatmap_data = df.pivot_table(index='Sentimen', columns='Emotion', aggfunc='size', fill_value=0)
                fig_heatmap = px.imshow(heatmap_data, text_auto=True, title='Sentiment vs Emotion Heatmap')
                st.plotly_chart(fig_heatmap)

    with tab2:
        kd = get_kamus_data()

        # for item in stqdm(kd, desc="Mengalisa data semantik konten ...", backend=False, frontend=True):
            # df[item["kamus_data"]] = df['Konten'].progress_apply(lambda x: classify_job_category2(str(x), item))


        formatted_array = [item["kamus_data"] for item in kd]
        st.dataframe(df[['Konten', 'Url'] + formatted_array])
        
        col1, col2 = st.columns(2, gap="large")

        with col1:
            for item in formatted_array:
                category_count = df[item].explode().value_counts().reset_index()
                category_count.columns = [item, 'Count']
                category_count_filtered = category_count[category_count[item] != 'tdk_ada_informasi']

                # Pie chart
                fig_pie_category = px.pie(category_count_filtered, names=item, values='Count', title=f'{item} Distribution in Captions')
                st.plotly_chart(fig_pie_category)

        with col2:
            for item in formatted_array:
                category_count = df[item].explode().value_counts().reset_index()
                category_count.columns = [item, 'Count']
                category_count_filtered = category_count[category_count[item] != 'tdk_ada_informasi']
                
                # Horizontal bar 
                fig_bar_category = px.bar(category_count_filtered, x='Count', y=item, orientation='h', title=f'{item} Distribution (Horizontal Bar Chart)')
                st.plotly_chart(fig_bar_category)

if not st.session_state.authentication_status:
    st.markdown("""
        <style>
            .stMainMenu{
                display: none !important;
            }
        </style>
    """, unsafe_allow_html=True)

    st.write("# Social Media Data Analysist 📈🚀")
    st.subheader("Dashboard for Insight & Semantic Data Analysist")

    col1, col2 = st.columns([2.5,1], vertical_alignment="center", gap="large")
    with col1:
        image_login = Image.open('login_analis.jpg')
        st.image(image_login, use_column_width=True)
    with col2:
        try: 
            authenticator.login(fields=dict({'Form name':'Silahkan Login 🔒😎', 'Username':'Masukan Username :', 'Password':'Masukan Password :', 'Login':'Lanjukan', 'Captcha':'Masukan Kode di bawah :'}), captcha=True)
        except Exception as e:
            st.toast(e)
else:
    # st.title(" Data Analysis for Pasar Kerja")
    st.write("# Mesin Analisa Klasifikasi Text 📈🚀")
    st.write("##### Identifikasi dan Klasifikasi Data Text tidak ter-struktur dengan akurat & presisi.")
    # st.write("-- --")

    st.markdown("""
        <style>
            div[data-testid='stNumberInputContainer']{
                border: 5px solid #f0f2f6;
            }
                
            div[data-testid='stNumberInputContainer'] input{
                background: white;
            }

            #editorCanvas > div.node-editor > .background{
                background-color: rgb(113, 113, 113) !important;
            }
        </style>
    """, unsafe_allow_html=True)

    ps.setup_st_sidebar(st, authenticator)

    client = init_connection()
    db = client.medmon

    items = db.config.find()
    arr = list(items)
    config = arr[0] if len(arr) > 0 else None

    image = Image.open('ilustrasi_old.png')
    

    # logo = st.image("logo.gif", caption="Sunrise by the mountains")
    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.image(image)

        st.text("")
        st.text("")

        with st.container(border=True):
            st.subheader('Klasifikasi Akun :')
            
            st.write('Kategori Akun Loker Mengandung Kata berikut :')
            
            if config:
                data_default_akun_loker = config["kat_akun_loker"]
            else:
                data_default_akun_loker = ['loker', 'karir', 'kerja']

            kat_akun_loker = st_tags(
                label="",
                text='Press enter to add more',
                value=data_default_akun_loker,
                maxtags = 20,
                key='1')
            
        with st.container(border=True):
            st.subheader('Preceding & Succeeding :')
            st.write("Berikut klasifikasi berdasarkan awalan dan akhiran :")
            tab1, tab2 = st.tabs(["Rentang Gaji", "Kouta Lowongan"])

            with tab1:
                pre_gaji = st_tags(
                    label="Kalimat Awalan :",
                    text='Press enter to add more',
                    value=["test 1", "test 2"],
                    maxtags = 20,
                    key='pre_gaji')
                
                succ_gaji = st_tags(
                    label="Kalimat Akhiran :",
                    text='Press enter to add more',
                    value=["test 1", "test 2"],
                    maxtags = 20,
                    key='succ_gaji')

    with col2:
        with st.container(border=True):
            st.subheader('Klasifikasi Utama :')
            st.write("Berikut klasifikasi berdasarkan Kamus Data.")

            tab1, tab2 = st.tabs(["Lowongan", "Non-Lowongan"])

            with tab1:
                workbook = load_workbook('kamus_data.xlsx', read_only=True)

                visible_sheets = [sheet for sheet in workbook.sheetnames if workbook[sheet].sheet_state == 'visible']

                if config:
                    data_default_keywords = config["keywords"]
                else:
                    data_default_keywords = visible_sheets

                keywords = st.multiselect(
                    "Kamus Data Analisa Semantik terkait Lowongan :",
                    options=visible_sheets,
                    default=data_default_keywords
                )

                
                left, right = st.columns(2)

                with right:
                    popover = st.popover("Opsi Kamus Data", icon=":material/settings:", use_container_width=True)

                    with open("kamus_data.xlsx", "rb") as file:
                        popover.download_button(
                            label="Unduh Kamus Data",
                            data=file,
                            file_name="kamus_data.xlsx",
                            use_container_width=True,
                            icon=":material/download:",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

                    if popover.button("Ganti Kamus Data",use_container_width=True,icon=":material/sync:"):
                        update_kamus_data()
                
                with st.expander("Bobot Klasifikasi", icon=":material/percent:"):
                    st.write('#### Bobot Klasifikasi (%)')
                    
                    st.write('Bobot ini menentukan nilai apakah suatu konten Lowongan / Non-Lowongan. Pastikan total dari semua bobot tidak lebih dari 100%.')

                    left, center, right = st.columns(3)

                    list_bobot = ["Akun Loker"] + keywords

                    for index, item in enumerate(list_bobot):
                        bobot_awal = 0

                        if(index==0):
                            bobot_awal = 50.00
                        else:
                            bobot_awal = 50/(len(list_bobot)-1)

                        if config and len(list_bobot) == len(config["nilai_bobot"]):
                            bobot_awal = config["nilai_bobot"][index]

                        if index % 3 == 0:
                            with left:
                                nilai_bobot.append(st.number_input(f"{item} :", value=bobot_awal))
                        elif index % 3 == 1:
                            with center:
                                nilai_bobot.append(st.number_input(f"{item} :", value=bobot_awal))
                        else:
                            with right:
                                nilai_bobot.append(st.number_input(f"{item} :", value=bobot_awal))
                    
                    total = sum(nilai_bobot)

                    if round(total, 2) < 100:
                        st.write(f":red[*Data Bobot Kurang dari 100%. Tambahkan nilai Bobot sebanyak {abs(round(100 - total, 2))} %.]")
                    elif round(total, 2) > 100:
                        st.write(f":red[*Data Bobot Lebih dari 100%. Kurangi nilai Bobot sebanyak {abs(round(100 - total, 2))} %.]")
                    else:
                        st.write(":green[*Total Data Bobot Sudah PAS 100%]")

            with tab2:
                st.write("Mohon Maaf, Klasifikasi Non-Lowongan masih tahap Pengembamgan.")


        with st.container(border=True):
            st.header('Proses Klasifikasi :')

            data_source = st.radio("Data Source to Analyze :", ["***Local (Excel)***", "***Remote (OP Server)***", "***Remote (WL Server)***"], horizontal=True)

            is_excel = data_source=="***Local (Excel)***"
            is_server = data_source=="***Remote (OP Server)***"
            
            if(is_excel):
                uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"])
            elif(is_server):
                st.write("Not Implemented")
            else:
                # Date slider range component
                date_range = st.date_input('Start Date  - End Date :', [])
                
                if(len(date_range) > 1):
                    from_date, to_date = date_range
                else:
                    from_date, to_date = [datetime.now(), datetime.now()]

            data_target = st.radio("Data Target Hasil Proses :", ["***Buat Baru (Create)***", "***Sudah Ada (Update)***"], horizontal=True)
                
            chunksize = st.slider(
                "Total Row Per Process :",
                value=500,
                min_value=100,
                max_value=2000,
                step=100
            )

            left, middle, right = st.columns(3)
            
            if(is_excel is not True and is_server is not True):
                with right:
                    if st.button("Export Data", type="secondary", icon="📥", use_container_width=True):
                        base_url = "https://api.kurasi.media/new-export/456"

                        dynamic_url = generate_url(base_url, from_date.strftime("%Y-%m-%d"), to_date.strftime("%Y-%m-%d"))

                        show_dynamic_url(dynamic_url)
        
            st.text("")

            left, right = st.columns(2) 

            with left:
                if st.button("Simpan Konfigurasi", type="secondary", icon=":material/save:", use_container_width=True):
                    
                    collection = db["config"]
                    
                    data_config = {
                        "keywords": keywords,
                        "kat_akun_loker": kat_akun_loker,
                        "nilai_bobot": nilai_bobot
                    }

                    collection.drop()

                    collection.insert_one(data_config)

                    st.toast("Berhasil Menyimpan Konfigurasi!")

            with right:
                eksekusi = st.button("Mulai Proses Data", type="primary", icon="▶️", use_container_width=True)

            
            # st.header('Result Proses :')

            # tab1, tab2 = st.tabs(["Insights Data Analysis", "Semantic Data Analysis"])
            st.text("")
            st.text("")

            if(eksekusi):
                st.toast("Memulai proses")
                if is_excel and uploaded_file is not None:
                    # Read the Excel file in chunks
                    # processed_chunks = []

                    with st.spinner('Memproses file excel ...'):
                        kd = get_kamus_data()
                        total_rows = pd.read_excel(uploaded_file, engine='openpyxl', sheet_name="Media Sosial").shape[0]
                        num_chunks = (total_rows // chunksize) + (total_rows % chunksize > 0)


                    progress_bar = st.progress(0, f"Memproses {num_chunks} Chunk Data.")
                    sub_progress_bar = st.progress(0, f"Menampilkan sub process ...")

                    current_time = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
                    # collection_name = f"hasil_proses_{current_time}"
                    collection_name = "hasil_proses_v2"

                    for i in range(num_chunks):
                        # Update the progress bar

                        # Use skiprows to read specific chunks
                        chunk = pd.read_excel(
                            uploaded_file,
                            sheet_name="Media Sosial",
                            engine='openpyxl',
                            skiprows=range(1, i * chunksize + 1),  # Skip rows that have already been processed
                            nrows=chunksize,  # Read only 'chunksize' rows at a time
                        )
                        
                        # Process the current chunk
                        # processed_chunk = process_chunk(chunk, kd)
                        # processed_chunks.append(processed_chunk)
                        
                        progress_bar.progress((i + 1) / num_chunks, f"Memproses {i + 1}/{num_chunks} chunk data")

                        process_chunk(chunk, kd, collection_name, sub_progress_bar)
                        
                        
                        progress_bar.progress((i + 1) / num_chunks, f"Mengirim data ke server {i + 1}/{num_chunks} ...")
                        time.sleep(0.5)  # Pause for 2 seconds between chunks
                    
                    # Combine all processed chunks into a single DataFrame
                    
                    progress_bar.empty()

                    st.toast("Sukses proses file", icon="ℹ️")
                    
                    # final_df = pd.concat(processed_chunks, ignore_index=True)
                    # eksekusi_excel2(final_df, kd)
                    
                    # eksekusi_excel(tab1, tab2, df)
                elif is_excel is not True and uploaded_file is None:
                    base_url = "https://api.kurasi.media/new-export/456"
                    dynamic_url = generate_url(base_url, from_date.strftime("%Y-%m-%d"), to_date.strftime("%Y-%m-%d"))

                    try:
                        with st.spinner('Mengambil data ke Server, mohon tunggu sebentar ...'):
                            response = requests.get(dynamic_url)
                            response.raise_for_status()

                            file_bytes = BytesIO(response.content)
                            df = pd.read_excel(file_bytes, sheet_name="Media Sosial")

                        # eksekusi_excel(tab1, tab2, df)

                    except requests.exceptions.RequestException as e:
                        st.error(f"Error: {e}")
                else:
                    st.toast("Ooppss, There's Something Wrong!", icon="ℹ️")
