import streamlit as st
import pandas as pd
import plotly.express as px
from PIL import Image
import re
import requests
from openpyxl import load_workbook
# from stqdm import stqdm
from io import BytesIO
from datetime import datetime
import time
import pymongo
from pymongo import UpdateOne
import streamlit_authenticator as stauth
import yaml
from yaml.loader import SafeLoader
import page_service as ps
import config_klasifikasi as ck
from streamlit_tags import st_tags
import json

# setup page
st.set_page_config(
    page_title="Social Media Data Analysist",
    page_icon="📊",
    layout="wide"
)
ps.setup_style_awal(st)
img_logo = Image.open('logo_pasker.png')
st.logo(img_logo, size="large")

# Check Authehntikasi
with open('config.yaml') as file:
    config_yaml = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config_yaml['credentials'],
    config_yaml['cookie']['name'],
    config_yaml['cookie']['key'],
    config_yaml['cookie']['expiry_days'],
    auto_hash=False
)
authenticator.login(location="unrendered")
st.session_state["auth_obj"] = authenticator
is_superadmin = False if st.session_state["name"] == "Analis Lowongan" else True

# Variable db remote
client = None
db = None

# Global Variable
dynamic_url = None
job_categories = None
uploaded_file = None
collection_name = ""
final_df = pd.DataFrame()

# Variable2 Step Klasifikasi Berurutan
kat_akun_loker = None
config_klasifikasi = []
keywords = []
nilai_rentang_gaji = {
    "pre": [],
    "succ": []
}
nilai_kouta = {
    "pre": [],
    "succ": []
}
nilai_bobot = []
selected_klasifikasi_aktif = []
indonesian_numbers = [
    "satu", "dua", "tiga", "empat", "lima", "enam", "tujuh", "delapan", 
    "sembilan", "sepuluh", "sebelas", "dua belas", "tiga belas", 
    "empat belas", "lima belas", "enam belas", "tujuh belas", 
    "delapan belas", "sembilan belas", "dua puluh", "tiga puluh", 
    "empat puluh", "lima puluh", "seratus", "seribu"
]
word_to_num = {
    'rb': 1000,
    'ribu': 1000,
    'jt': 1_000_000,
    'juta': 1_000_000,
    'yen': 100,
    'm': 1_000_000,
}

# Function2 Klasifikasi
def classify_job_category(caption, categories):
    matched_categories = []

    # Iterate through each category
    for category in categories["category_data"]:
        # Check if any threshold word is present in the caption (ignore case)
        for keyword in category["threshold"]:
            if re.search(r"\b" + re.escape(keyword) + r"\b", caption, re.IGNORECASE):
            # if re.search(r"(?<!\w)" + re.escape(keyword) + r"(?!\w)", caption, re.IGNORECASE):
            # if re.search(r"(?<![a-zA-Z])" + re.escape(keyword) + r"(?![a-zA-Z])", caption, re.IGNORECASE):
                matched_categories.append(category["name"])
                break  # Move to the next category once a match is found

    if not matched_categories:
        matched_categories.append("tdk_ada_informasi")

    if "tdk_ada_informasi" in matched_categories and len(matched_categories) > 1:
        matched_categories.remove("tdk_ada_informasi")

    return matched_categories

def classify_akun(nama_akun, is_pers_value):
    kat_akun = "Akun Netizen"
    pattern = re.compile(r'|'.join(kat_akun_loker), re.IGNORECASE)

    if(is_pers_value):
        return "Akun Pers"
    
    if pattern.search(nama_akun):
        return "Akun Loker"
    
    return kat_akun

def kalkulasi_persentase_lowongan(params):
    result = 0
    params.insert(0, params.pop())

    for i, param in enumerate(params):
        if(i==0 and param=="Akun Loker"):
            result += nilai_bobot[0]
        elif(i != 0 and "tdk_ada_informasi" not in param):
            result += nilai_bobot[i]

    return result

def classify_proceeding_succeeding(caption, pro_suc):
    pre_list = pro_suc['pre']
    succ_list = pro_suc['succ']
    
    # Split the caption into sentences using common sentence delimiters
    # sentences = re.split(r'[.!?\n]', caption)
    
    pre_pattern = r"|".join(map(re.escape, pre_list))
    succ_pattern = r"|".join(map(re.escape, succ_list))


    pattern = (
        r"(?i)\b(?:"
        + pre_pattern
        + r")\b"                          # Matches any item in 'pre' list as a whole word
        + r"(?:\s+\w+){0,5}\s*"               # Optionally match intermediate words
        + r"([\d.,]+(?:\s*(?:sampai|-)\s*[\d.,]+)?)"  # Capture numbers or ranges
        + r"(?:\s*(?:" + succ_pattern + r"))?\b"       # Match any item in 'succ' list as optional
    )

    match = re.search(pattern, caption, re.IGNORECASE)
    
    if match:
        # matched_str = match.group(0)
        # result = re.sub(r"(?i)\bgaji\b", "", matched_str).strip()

        result = match.group(0)
        
    else: 
        result = ""

    return result

    # Return an empty string if no valid match is found
    # return ""

# Funciton2 Flow Pemrosesan
def process_chunk(chunk, kd, progress_bar):
    for index, row in chunk.iterrows():
        # Get the 'Konten' value once for this row
        konten_value = str(row['Konten'])

        if "Akun" in selected_klasifikasi_aktif:
            akun_value = str(row['Akun/Judul'])
            is_pers_value = str(row['Jenis Akun']) == "Pers"
        
        data_params_kalkulasi_persentase_lowongan = []
        
        if "Lowongan" in selected_klasifikasi_aktif:
            # Calculate and set classifications for each item in kd
            for item in kd:
                # Assuming item["kamus_data"] is the name of the new column
                column_name = item["kamus_data"]

                if column_name not in chunk.columns:
                    # Create the column if it doesn't exist
                    chunk[column_name] = None
                
                # Apply classify_job_category and assign result to the specific cell
                result_1 = classify_job_category(konten_value, item)
                chunk.at[index, column_name] = result_1
                data_params_kalkulasi_persentase_lowongan.append(result_1)

        if "Rentang Gaji" in selected_klasifikasi_aktif:
            # Klasifikasikan Rentang Gaji by Succeding & Proceeding
            if "Rentang Gaji" not in chunk.columns:
                chunk["Rentang Gaji"] = None

            if "Digit Gaji (Clean)" not in chunk.columns:
                chunk["Digit Gaji (Clean)"] = None

            result_3 = classify_proceeding_succeeding(konten_value, nilai_rentang_gaji)
            chunk.at[index, "Rentang Gaji"] = result_3
            if result_3:
                chunk.at[index, "Digit Gaji (Clean)"] = extract_salary(result_3)

            data_params_kalkulasi_persentase_lowongan.append(result_3)
            
        if "Kouta Lowongan" in selected_klasifikasi_aktif:
            # Klasifikasikan Kouta Lowongan by Succeding & Proceeding
            if "Kouta Lowongan" not in chunk.columns:
                chunk["Kouta Lowongan"] = None

            if "Digit Kouta (Clean)" not in chunk.columns:
                chunk["Digit Kouta (Clean)"] = None

            result_4 = classify_proceeding_succeeding(konten_value, nilai_kouta)
            chunk.at[index, "Kouta Lowongan"] = result_4
            
            if result_4:
                chunk.at[index, "Digit Kouta (Clean)"] = extract_quota(result_4)
            else:
                chunk.at[index, "Digit Kouta (Clean)"] = 1

            data_params_kalkulasi_persentase_lowongan.append(result_4)

        if "Akun" in selected_klasifikasi_aktif:
            # Klasifikasikan By Pattern Tipe Akun
            if "Klasifikasi Akun" not in chunk.columns:
                chunk["Klasifikasi Akun"] = None
            result_2 = classify_akun(akun_value, is_pers_value)
            chunk.at[index, "Klasifikasi Akun"] = result_2
            data_params_kalkulasi_persentase_lowongan.append(result_2)

        if "Bobot Lowongan" in selected_klasifikasi_aktif:
            # Klasifikasikan Bobot Menghitung Persentase Lowongan
            if "Persentase Lowongan" not in chunk.columns:
                chunk["Persentase Lowongan"] = None

            chunk.at[index, "Persentase Lowongan"] = kalkulasi_persentase_lowongan(data_params_kalkulasi_persentase_lowongan)

        progress_value = min(max(index/len(chunk), 0.0), 1.0)
        progress_bar.progress(progress_value, f"Menganalisa data baris ke-{index} dari {len(chunk)} data.")

    progress_bar.progress(0, "Progress Indentifikasi & Klasifikasi Text ...")

    return chunk

def get_kamus_data():
    kamus_data_xls = pd.ExcelFile('kamus_data.xlsx')
    result = []
    
    for sheet_name in keywords:
        # Load the sheet into a DataFrame
        df = pd.read_excel(kamus_data_xls, sheet_name=sheet_name)

        # Remove columns that are completely empty (all NaN values)
        df = df.dropna(axis=1, how='all')
        
        # Remove rows that are completely empty (all NaN values)
        df = df.dropna(axis=0, how='all')
        
        
        # Initialize a list to hold column data for this sheet
        column_data = []

        # Iterate through each column in the DataFrame
        for column in df.columns:
            # Get the column data as a list
            list_data = df[column].dropna().tolist()
            
            # Add column name and list of data to column_data list
            if list_data:
                column_data.append({
                    "name": column,
                    "threshold": list_data
                })
            
        # Append the sheet data to result
        if column_data:
            result.append({
                "kamus_data": sheet_name,
                "category_data": column_data
            })

    return result


# Funtion2 Utilitas
def contains_number_or_word(sentence):
    # Check if the sentence contains any digit
    if re.search(r'\d', sentence):
        return True
    # Check if the sentence contains any number word from the list
    for word in indonesian_numbers:
        if re.search(r'\b' + re.escape(word) + r'\b', sentence, re.IGNORECASE):
            return True
    return False

def clean_number(num_str):
    # Ganti tanda koma menjadi titik jika digunakan sebagai desimal
    num_str = num_str.replace(',', '.')
    # Hapus titik yang berfungsi sebagai pemisah ribuan
    if num_str.count('.') > 1 or ('.' in num_str and len(num_str.split('.')[-1]) > 2):
        num_str = num_str.replace('.', '')
    return num_str

def extract_salary(text):
    # Normalisasi teks ke huruf kecil
    text = text.lower()

    # Regex untuk menemukan angka (dengan atau tanpa desimal)
    matches = re.findall(r'(\d+[\.,]?\d*)\s*(rb|ribu|jt|juta|yen|m)?', text)

    # Jika tidak ada angka yang ditemukan, return 0
    if not matches:
        return 0

    # Konversi angka ke nominal
    salaries = []
    for match in matches:
        num_str, unit = match
        num_str = clean_number(num_str)
        num = float(num_str)

        # Jika unit ditemukan, konversikan ke nilai
        if unit:
            multiplier = word_to_num.get(unit, 1)
            num *= multiplier
        else:
            # Jika tidak ada satuan dan angka di bawah 1000, anggap itu ratusan ribu
            if num < 1000:
                num *= 1000

        salaries.append(num)

    # Jika ada range atau lebih dari satu gaji, cari rata-rata
    if len(salaries) > 1:
        avg_salary = sum(salaries) / len(salaries)
    else:
        avg_salary = salaries[0]

    # Periksa konteks "per hari" atau "per jam"
    if 'per hari' in text or '/ hari' in text:
        avg_salary *= 22  # Anggap 22 hari dalam sebulan
    elif 'per jam' in text or '/ jam' in text:
        avg_salary *= 8 * 22  # Anggap 8 jam kerja per hari dan 22 hari kerja per bulan

    if int(avg_salary) < 100000:
        return None
    else: 
        return int(avg_salary)

def extract_quota(text):
    # Normalisasi teks ke huruf kecil
    text = text.lower()

    # Regex untuk menemukan angka (dengan atau tanpa pemisah ribuan)
    matches = re.findall(r'(\d+[\.,]?\d*)', text)

    # Jika tidak ada angka yang ditemukan, return 0
    if not matches:
        return 0

    # Bersihkan angka dari tanda pemisah ribuan dan konversikan ke integer
    quotas = []
    for num_str in matches:
        num_str = num_str.replace('.', '').replace(',', '')
        num = int(num_str)
        quotas.append(num)

    # Ambil angka pertama yang ditemukan (karena biasanya kuota hanya satu angka yang dominan)
    return quotas[0] if quotas else 1

@st.cache_resource
def init_connection():
    return pymongo.MongoClient(**st.secrets["mongo"])

def generate_url(base_url, from_date, to_date, date_type="date", sort="date"):
    source = "ZmFjZWJvb2ssdHdpdHRlcixpbnN0YWdyYW0sdGlrdG9rLHlvdXR1YmUsZm9ydW0sYmxvZyxsaW5rZWRpbg=="
    now = datetime.now()
    cur_time = now.strftime("%Y%m%d%H%M%S")

    return f"{base_url}?from={from_date}%2000:00:00&to={to_date}%2023:59:59&date_type={date_type}&sort={sort}&sources={source}&time={cur_time}"

def convert_df_to_excel(df):
    # Create a copy of the DataFrame to avoid modifying the original
    df_copy = df.copy()
    
    # Convert any column with list values to comma-separated strings
    for col in df_copy.columns:
        df_copy[col] = df_copy[col].apply(lambda x: ', '.join(x) if isinstance(x, list) else x)
    
    # Write the DataFrame to an Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_copy.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)
    return output

def get_collections_with_prefix(prefix):
    collections = db.list_collection_names()
    filtered_collections = [coll for coll in collections if coll.startswith(prefix)]
    return filtered_collections

def execute_query_builder(db_source, fields, sort_fields, sort_order_value):
    result = None
    
    try:
        collection = db[db_source]
        # Parse the query input as JSON
        query = json.loads(query_input)

        # Construct projection for selected fields
        projection = {field: 1 for field in fields} if fields else None

        # Construct sort criteria
        sort_criteria = [(field, sort_order_value) for field in sort_fields] if sort_fields else None

        # Execute the query
        cursor = collection.find(query, projection)
        if sort_criteria:
            cursor = cursor.sort(sort_criteria)
        results_list = list(cursor.limit(total_limit))

        result = results_list

    except json.JSONDecodeError:
        result = {"msg": "Error."}

    return result

def parse_json_telegram(json_data, keywords):
    messages = json_data.get("messages", [])
    follower = json_data.get("follower")
    rows = []

    for msg in messages:
        content = "".join(
            [t.get("text", "") if isinstance(t, dict) else t for t in msg.get("text", [])]
        )

         # Filter messages based on keywords
        if not any(keyword.lower() in content.lower() for keyword in keywords):
            continue  # Skip messages that don't contain the keywords

        # Extract required fields
        date = msg.get("date", None)
        time_part = date.split("T")[1] if date else None

        account_name_1 = msg.get("from")
        account_name_2 = msg.get("from_id")
        
        account_name = account_name_1 if account_name_1 else account_name_2

        urls = [entity.get("text") for entity in msg.get("text_entities", []) if entity.get("type") == "link"]
        url = urls[0] if urls else None

        # Static values
        topik = "Lowongan"
        sumber = "Telegram"

        # Placeholder values for Likes, View, Engagement (not in the provided JSON)
        likes = None
        view = None
        engagement = None

        # Append row
        rows.append({
            "Tanggal Publikasi": date.split("T")[0] if date else None,
            "Jam Publikasi": time_part if time_part else None,
            "Tanggal Tersimpan": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Topik": topik,
            "Akun/Judul": account_name,
            "Konten": content,
            "Sumber": sumber,
            "Url": url,
            "Likes": likes,
            "View": view,
            "Engagement": engagement,
            "Jenis Akun": "Non Pers",
            "Followers": follower
        })

    # Create DataFrame
    return pd.DataFrame(rows)

# Fungsi-fungsi Konfig Klasifikasi
def init_config_klasifikasi(arrModeKlasifikasi):
    for mode in arrModeKlasifikasi:
        config_klasifikasi.append(mode)


def konfig_klasifikasi_by_kamus_data(idx):
    st.write("test")

# Modal & Dialog
@st.dialog("Download disini :")
def show_dynamic_url(url):
    st.write(f"{url}")

@st.dialog("Hasil Proses Klasifikasi :", width="large")
def preview_result(kd, data_df, tipe=None):
    # formatted_array = [item["kamus_data"] for item in kd]
    # data_df = data[['Akun/Judul','Konten', 'Url', 'Rentang Gaji', 'Kouta Lowongan']]
    # data_df = data[['Akun/Judul','Konten', 'Url'] + formatted_array]

    if(tipe=="json"):
        st.write("#### Hasil Proses :")
        
        with st.container(border=True):
            st.json(data_df.to_json(orient='records', lines=True), expanded=False)
    else:
        st.dataframe(data_df)

        col = st.columns([3,1], gap="large")
        with col[1]:
            st.download_button(
                label="Download",
                use_container_width=True,
                type="primary",
                icon=":material/download:",
                data=convert_df_to_excel(data_df),
                file_name='hasil_proses.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

@st.dialog("Upload New Data :")
def update_kamus_data():
    data_baru = st.file_uploader("Upload Kamus Data :", type=["xlsx", "xls"])

    left, middle, right = st.columns(3)

    with right:
       if st.button("Lanjutkan", use_container_width=True, disabled=not data_baru):
           # Save the uploaded file to replace the existing one
            with open("kamus_data.xlsx", "wb") as f:
                f.write(data_baru.getbuffer())
                st.rerun()


if not st.session_state.authentication_status:
    st.markdown("""
        <style>
            .stFormSubmitButton {
                text-align: end;
            }
        </style>
    """, unsafe_allow_html=True)

    cols = st.columns([2,1], vertical_alignment="top", gap="large")
    with cols[0]:
        st.write("# Social Media Data Analysist 📈🚀")
        st.subheader("Dashboard for Insight & Semantic Data Analysist")

        image_login = Image.open('login_analisis_new.webp')
        st.image(image_login, width=450)
    with cols[1]:
        try: 
            authenticator.login(fields=dict({'Form name':'Silahkan Login 🔒😎', 'Username':'Masukan Username :', 'Password':'Masukan Password :', 'Login':'🔑 Akses Dashboard', 'Captcha':'Masukan Kode di bawah (Captcha) :'}), captcha=True)
        except Exception as e:
            st.toast(e)
else:
    _ = """ col = st.columns([4.5,1.3,1], vertical_alignment="center")

    with col[0]:
        st.write("## Mesin Klasifikasi Text 📈🚀")
        st.write("###### Identifikasi dan Klasifikasi Data Text tidak ter-struktur dengan akurat & presisi.")
    with col[1]:
        popover = st.popover("Filter Metode", icon=":material/filter_alt:", use_container_width=True)
        popover.checkbox("By Kamus Data", True)
        popover.checkbox("By Pattern", True)
        popover.checkbox("By Preceding & Succeeding", True)
        popover.checkbox("Bobot Klasifikasi", True)
    with col[2]:
        st.button("Guideline", icon=":material/info:", use_container_width=True) """

    st.markdown("""
        <style>
            div[data-testid='stNumberInputContainer']{
                border: 5px solid #f0f2f6;
            }
                
            div[data-testid='stNumberInputContainer'] input{
                background: white;
            }

            #editorCanvas > div.node-editor > .background{
                background-color: rgb(113, 113, 113) !important;
            }
                
            div.stMainBlockContainer.block-container > div > div > div > div.stHorizontalBlock > div.stColumn:nth-child(2) > div > div > div > div[data-testid='stVerticalBlockBorderWrapper']{
                border: 3px solid #666566;
                # background: #fff9f7;
            }
                
            .stMainBlockContainer.block-container > div > div > div > div > div.stColumn:nth-child(1) > div{
                top: 90px;
                position: sticky;
            }
        </style>
    """, unsafe_allow_html=True)

    st.markdown("""
        <style>
            .box-test {
                background: #666566;
                height: 3.5px;
                position: relative;
                overflow: visible;
                transform: translate(calc(-100% - 15px), 45px);
                width: 40px;
            }
                
            .box-test.add{
                top: -20px;
            }
                    
            .box-test > div.line{
                position: absolute;
                left: 0px;
                top: 0px;
                height: 2000px;
                width: 2px;
            }
                    
            .box-test > div.line.white{
                background: white;
                transform: translateX(-1px);
                width: 4px;
            }
                    
            .box-test > div.line.black{
                background: #666566;
            }
                    
            .stMainBlockContainer > div{
                contain: paint;
            }
                
            .box-test > div.step{
                position: absolute;
                top: 50%;
                transform: translate(-50%, -50%);
                background: #666566;
                color: white;
                width: 35px;
                height: 35px;
                border-radius: 100%;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 21px;
                font-weight: 900;
            }
        </style>
    """, unsafe_allow_html=True)

    ps.setup_st_sidebar(st, authenticator)

    client = init_connection()
    db = client.medmon

    items = db.config.find()
    arr = list(items)
    config = arr[0] if len(arr) > 0 else None

    col1, col2 = st.columns([1.5,2], gap="large")

    with col2:
        st.write("### Urutan Alur Kerja Klasifikasi :")
        list_collection = get_collections_with_prefix("hasil_proses_")

        # Sumber Klasifikasi
        with st.container(border=True):
            st.html(f"""
                <div class="box-test">
                    <div class="line black"></div>
                    <div class="step">1</div>
                </div>
            """)
           
            
            col = st.columns([2,1.2,1.2], vertical_alignment="bottom")
            
            with col[0]:
                st.subheader('Sumber Text (Input)')
            with col[1]:
                mode_klasifikasi = st.selectbox("", ["Proses Baru", "Proses Ulang"], label_visibility="collapsed")
            # Konfig Sumber
            with col[2]:
                config_source = st.popover("Konfig", icon=":material/settings:", use_container_width=True)

            
            config_source.write("###### Total Row Per Process :")

            chunksize = config_source.slider(
                "",
                value=500,
                min_value=100,
                max_value=2000,
                step=100,
                label_visibility="collapsed"
            )

            config_source.write("###### Aktifkan Klasifikasi :")
            options = ["Lowongan", "Non-Lowongan", "Akun", "Bobot Lowongan", "Rentang Gaji", "Kouta Lowongan"]

            if config and "selected_klasifikasi_aktif" in config:
                data_default_selected_klasifikasi_aktif = config["selected_klasifikasi_aktif"]
            else:
                data_default_selected_klasifikasi_aktif = options
            
            for idx, option in enumerate(options):
                val_ska = True if option in data_default_selected_klasifikasi_aktif else False

                if config_source.checkbox(option, value=val_ska):
                    selected_klasifikasi_aktif.append(option)

            
            if mode_klasifikasi=="Proses Baru":
                # st.write('Tentukan Sumber Text yang ingin di klasifikasi :')

                data_source = st.radio("Tipe Source :", ["Data Lokal", "Data Server", "Masukan Manual", "Export Telegram"], horizontal=True)
                
                match data_source:
                    case "Data Lokal":
                        uploaded_files = st.file_uploader("Upload Excel files", type=["xlsx", "xls"], accept_multiple_files=True)
                    case "Data Server":
                        col = st.columns([3.1,1], vertical_alignment="bottom")
                        
                        with col[0]:
                            date_range = st.date_input('Start Date  - End Date :', [])
                    
                            if(len(date_range) > 1):
                                from_date, to_date = date_range
                            else:
                                from_date, to_date = [datetime.now(), datetime.now()]
                        with col[1]:
                            base_url = "https://api.kurasi.media/new-export/456"
                            dynamic_url = generate_url(base_url, from_date.strftime("%Y-%m-%d"), to_date.strftime("%Y-%m-%d"))

                            st.link_button(
                                label="Lihat Data",
                                url=dynamic_url,
                                use_container_width=True,
                                icon=":material/visibility:",
                                disabled=not date_range,
                                type="primary"
                            )
                    case "Masukan Manual":
                        data_text_manual = st.text_area(f"Masukan Text : ", placeholder=f"Masukan Text disini ... ")
                    case "Export Telegram":
                        keyword_filter_tele = st_tags(
                        label="Filter Data Chat Mengandung Kata :",
                        text='Press enter to add more',
                        value=["lowongan", "recruitment", "loker", "hiring", "rekrutmen"],
                        maxtags = 200,
                        key='filter_tele')
                        uploaded_files = st.file_uploader("Upload JSON Exported Files from Telegram :", type=["json"], accept_multiple_files=True)

                        col = st.columns(3)

                        with col[2]:
                            preview = st.button("Preview Data", use_container_width=True, type="primary", icon=":material/search:")
                        
                        if preview:
                            for uploaded_file in uploaded_files:
                                st.write(f"#### Preview Data : {uploaded_file.name}")
                                json_data = json.load(uploaded_file)
                                df = parse_json_telegram(json_data, keyword_filter_tele)

                                # Display the DataFrame
                                st.write(df)

                    case default:
                        st.error("Not Implemented", icon=":material/info:")

            else:
                col = st.columns(2)
                
                with col[0]:
                    db_source = st.selectbox("Pilih Tabel : ", list_collection)
                    list_kolom = db[db_source].find_one().keys() if db[db_source].find_one() else []
                    sort_fields = st.multiselect("Sort By:", list_kolom)

                with col[1]:
                    fields = st.multiselect("Select Data :", list_kolom, key="kolom_source")
                    sort_order = st.radio("Sort order:", ["Ascending", "Descending"], horizontal=True)
                    sort_order_value = 1 if sort_order == "Ascending" else -1

                col2 = st.columns([4,1])

                with col2[0]:
                    example_data = '{"$and":[{"Lokasi Kota":{"$exists":false}},{"Jabatan":{"$exists":false}}]}'
                    query_input = st.text_area("Enter your MongoDB query as JSON:", value=example_data)

                with col2[1]:
                    total_limit = st.number_input("Limit Data :", value=10)

                st.text("")
                st.text("")

                col3 = st.columns([3,1.3])

                with col3[1]:
                    test_query = st.button("Test Query", icon=":material/play_arrow:", use_container_width=True, type='primary')

                if test_query:
                    with st.container(border=True):
                        # Display results
                        st.write("Query Results:") 
                        result_list = execute_query_builder(db_source, fields, sort_fields, sort_order_value)
                        st.json(result_list, expanded=False)

        # Pilih Metode Klasifikasi
        with st.container(border=True, key="pil_metode"):
            st.html(f"""
                <div class="box-test">
                    <div class="line black"></div>
                    <div class="step">2</div>
                </div>
            """)

            st.markdown("""
                <style>
                    [class*="st-key-pil_metode"] [data-testid="stWidgetLabel"] p > span{
                        padding: 5px 10px !important;
                        font-size: 14px !important;
                        font-weight: 900;
                        border-radius: 10px !important;
                        zoom: 0.8;
                    }
                        
                    [class*="st-key-pil_metode"] [data-testid="stWidgetLabel"] p > span:nth-child(1){
                        background: #ffd951 !important;
                        color: black !important;
                    }
                        
                    [class*="st-key-pil_metode"] [data-testid="stWidgetLabel"] p > span:nth-child(2){
                        background: #6b6c6f !important;
                        color: white !important;
                    }
                        
                    @media (max-width: 767px) {
                        [class*="st-key-pil_metode"] div[data-baseweb='tab-list']{
                            zoom: 0.5 !important;
                        }
                    }

                    [class*="st-key-pil_metode"] div[data-baseweb='tab-list']{
                        gap: 0px;
                        justify-content: center;
                        zoom: 0.8;
                    }
                        
                    [class*="st-key-pil_metode"] button[data-baseweb='tab']{
                        padding: 0px 50px;
                        border: 2px solid gray;
                        border-radius: 10px;
                    }
                        
                    [class*="st-key-pil_metode"] button[data-baseweb='tab']:hover{
                        border: 2px solid red;    
                    }

                    [class*="st-key-pil_metode"] button[aria-selected='true']{
                        border: 2px solid red;
                        background: #ffeded;
                    } 
                    
                    [class*="st-key-pil_metode"] button[aria-selected='true'] *{
                        font-weight: 900;
                    }

                    [class*="st-key-pil_metode"] button[data-baseweb='tab']:nth-child(1){
                        border-top-right-radius: 0px;
                        border-bottom-right-radius: 0px;
                        # border-right: unset;
                    }
                    
                    [class*="st-key-pil_metode"] button[data-baseweb='tab']:nth-child(2){
                        border-radius: unset;
                    }
                        
                    [class*="st-key-pil_metode"] button[data-baseweb='tab']:nth-child(3){
                        border-top-left-radius: 0px;
                        border-bottom-left-radius: 0px;
                    }
                        
                    [class*="st-key-pil_metode"] div[data-baseweb='tab-highlight'], [class*="st-key-pil_metode"] div[data-baseweb='tab-border']{
                        display: none;    
                    }
                        
                    [class*="st-key-card_klasifikasi_"] div.stExpander > details{
                        border: unset !important;
                    }
                    
                    [class*="st-key-card_klasifikasi_"] div.stExpander summary{
                        padding: unset;
                    }
                        
                    [class*="st-key-card_klasifikasi_"] div.stExpander summary svg{
                        display: none;
                    }
                        
                    [class*="st-key-card_klasifikasi_"] div[data-testid="stExpanderDetails"]{
                        padding: 20px 10px;
                        background: #f4f4f4;
                        border-radius: 20px;
                    }
                </style>
            """, unsafe_allow_html=True)

            st.subheader('Metode Klasifikasi')

            # Read JSON data from a local file
            file_path = "metode_klasifikasi_2.json"  # Replace with your file path

            metode_klasifikasi = []
            
            with open(file_path, "r") as file:
                list_metode_klasifikasi = json.load(file)

            num_columns = 3
            
            tabs = st.tabs(["Pilih Metode", "Urutan Metode", "Detail Metode"])

            with tabs[0]:
                st.write('Pilih Metode Klasifikasi yang di pakai, dengan AI maupun Non-AI :')
                with st.container(border=False, height=300):
                    columns = st.columns(num_columns)
                    for index, method in enumerate(list_metode_klasifikasi):
                        column_index = index % num_columns

                        is_ai = f':blue[{method["AI/Non-AI"]}]' if method["AI/Non-AI"] == "With AI" else f':red[{method["AI/Non-AI"]}]'
                        
                        with columns[column_index]:
                            if st.checkbox(f'{method["Name of Method"]}\n\n:red[{method["Category"]}] {is_ai}', key=f"method_{index}", disabled=not method["Is Implemented"]):
                                metode_klasifikasi.append(method)
            with tabs[1]:
                st.write("Under Development")
            with tabs[2]:
                st.dataframe(list_metode_klasifikasi)
            # metode_klasifikasi = st.multiselect("", ["Klasifikasi By Pattern", "Klasifikasi By Kamus Data", "Klasifikasi By Proceeding & Succeeding", "Bobot Klasifikasi (%)"], label_visibility="collapsed", placeholder="Pilih Metode disini ...")

            # if st.button("Tambah Klasifikasi", use_container_width=True, icon=":material/add:")
            init_config_klasifikasi(metode_klasifikasi)

            
        
        # Klasifikasi Dinamis
        for idx, m_klasifikasi in enumerate(metode_klasifikasi):
            with st.container(border=True, key=f"card_klasifikasi_{idx}"):
                st.html(f"""
                    <div class="box-test">
                        <div class="line black"></div>
                        <div class="step">{idx+3}</div>
                    </div>
                """)

                col = st.columns([3,1.2], vertical_alignment="center")
                
                with col[0]:
                    st.subheader(f'{m_klasifikasi["Name of Method"]}')
                with col[1]:
                    config_kd = st.popover("Konfigurasi", icon=":material/tune:", use_container_width=True)

                config_kd.write("##### Daftar Klasifikasi Data :")

                with config_kd.container():
                    st.markdown("<div style='width: 400px;'></div>", unsafe_allow_html=True)
                    
                    if 'list_klasifikasi_data' not in config_klasifikasi[idx]:
                        config_klasifikasi[idx]["list_klasifikasi_data"] = [
                            {"Judul": "Klasifikasi 1", "Dari Kolom": "" }
                        ]

                        nd = pd.DataFrame(config_klasifikasi[idx]["list_klasifikasi_data"])
                        nd["Dari Kolom"] = nd["Dari Kolom"].astype("category")
                        nd["Dari Kolom"] = nd["Dari Kolom"].cat.add_categories(("☯ Neutral", "😤 Negative"))

                    df = st.data_editor(nd, num_rows="dynamic", key=f"de_{idx}", use_container_width=True, hide_index=True)
                    filtered_df = df[
                        df["Judul"].apply(lambda x: isinstance(x, str)) &
                        df["Dari Kolom"].apply(lambda x: isinstance(x, str))
                    ]
                    
                    config_klasifikasi[idx]["list_klasifikasi_data"] = filtered_df 

                if not config_klasifikasi[idx]["list_klasifikasi_data"].empty:

                    with st.expander(f'This Method {m_klasifikasi["Short Description"]}', expanded=False):
                        st.write(f'{m_klasifikasi["Long Description"]}')
                        st.write("##### Example Case :")
                        st.code(f'{m_klasifikasi["Example Case"]}')
            
                    judul_array = config_klasifikasi[idx]["list_klasifikasi_data"]["Judul"].tolist()
                    
                    tabs_kd = st.tabs(judul_array)

                    for idx2, tab_kd in enumerate(tabs_kd):
                        with tab_kd:
                            match m_klasifikasi["id"]:
                                case "BW-001":
                                    ck.konfig_klasifikasi_bw001(idx2,config, st_tags)
                                case "BW-004":
                                    ck.konfig_klasifikasi_bw004(idx2,config, st)
                                case "BW-006":
                                    ck.konfig_klasifikasi_bw006(idx2,config, st_tags)
                                case _:
                                    col = st.columns([1,3,1])
                                    with col[1]:
                                        st.image("https://www.rackh.com/wp-content/uploads/2023/06/18771510_6029646-800x400.jpg", use_container_width=True)
                                        st.write("### :red[Method Not Implemented]")

                                    st.caption(f'Metode :blue[{m_klasifikasi["Name of Method"]}] untuk klasifikasi teks saat ini masih dalam tahap pengembangan dan belum diimplementasikan. Proses pengembangan melibatkan penelitian mendalam, pengujian algoritma, serta integrasi ke dalam sistem yang ada untuk memastikan metode ini dapat memenuhi kebutuhan pengguna dengan optimal.')
                else:
                    st.write("Belum ada Klasifikasi Kamus Data.")

        # Preview Data Config
        st.json(config_klasifikasi, expanded=False)

        # Klasifikasi Akun by Pattern
        with st.container(border=True):
            st.html(f"""
                <div class="box-test">
                    <div class="line black"></div>
                    <div class="step">2</div>
                </div>
            """)

            st.subheader('Klasifikasi By Pattern')
            
            st.write('Kategori Akun Loker Mengandung Kata berikut :')
            
            if config:
                data_default_akun_loker = config["kat_akun_loker"]
            else:
                data_default_akun_loker = ['loker', 'karir', 'kerja']

            kat_akun_loker = st_tags(
                label="",
                text='Press enter to add more',
                value=data_default_akun_loker,
                maxtags = 200,
                key='1')
        
        # Klasifikasi Lowongan by Kamus Data
        with st.container(border=True):
            st.html(f"""
                <div class="box-test">
                    <div class="line black"></div>
                    <div class="step">3</div>
                </div>
            """)

            col = st.columns([3,1.2], vertical_alignment="bottom")
            
            with col[0]:
                st.subheader('Klasifikasi By Kamus Data')
            with col[1]:
                config_kd = st.popover("Kamus Data", icon=":material/settings:", use_container_width=True)

            with config_kd.container():
                list_kbkd = st_tags(
                    label="#### Masukan Kamus Data :",
                    text='Tekan Enter Untuk Tambah ...',
                    maxtags = 200,
                    value=["Lowongan", "Persepsi Netizen"],
                    key='data_kd'
                )

            # st.subheader('Klasifikasi By Kamus Data')
            if list_kbkd:
                st.write("Berikut klasifikasi berdasarkan Kamus Data.")

                tabs_kd = st.tabs(list_kbkd)

                for idx, tab_kd in enumerate(tabs_kd):
                    with tab_kd:
                        workbook = load_workbook('kamus_data.xlsx', read_only=True)

                        visible_sheets = [sheet for sheet in workbook.sheetnames if workbook[sheet].sheet_state == 'visible']

                        if config:
                            if config["keywords"] == visible_sheets:
                                data_default_keywords = config["keywords"]
                            else: 
                                data_default_keywords = visible_sheets
                        else:
                            data_default_keywords = visible_sheets

                        keywords = st.multiselect(
                            "Kamus Data Analisa Semantik terkait Lowongan :",
                            options=visible_sheets,
                            default=data_default_keywords,
                            key=f"keywords-{idx}"
                        )

                        
                        left, right = st.columns([2,1])

                        with right:
                            popover = st.popover("Opsi Kamus Data", icon=":material/settings:", use_container_width=True, disabled=not is_superadmin)

                            with open("kamus_data.xlsx", "rb") as file:
                                popover.download_button(
                                    label="Unduh Kamus Data",
                                    data=file,
                                    file_name="kamus_data.xlsx",
                                    use_container_width=True,
                                    icon=":material/download:",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"download-kd-{idx}"
                                )

                            if popover.button("Ganti Kamus Data",use_container_width=True,icon=":material/sync:", key=f"ganti-kd-{idx}"):
                                update_kamus_data()
            else:
                st.write("Belum ada Klasifikasi Kamus Data.")

        # Klasifikasi Rengang Gaji & Kouta Lowongan by Proceeding & Succeding
        with st.container(border=True):
            st.html(f"""
                <div class="box-test">
                    <div class="line black"></div>
                    <div class="step">4</div>
                </div>
            """)
            
            st.subheader('Klasifikasi By Proceeding & Succeeding')

            st.write("Berikut klasifikasi berdasarkan awalan dan akhiran :")
            tab1, tab2 = st.tabs(["Rentang Gaji", "Kouta Lowongan"])

            if config:
                data_default_nilai_rentang_gaji = config["nilai_rentang_gaji"]
                data_default_nilai_kouta = config["nilai_kouta"]
            else:
                data_default_nilai_rentang_gaji = nilai_rentang_gaji
                data_default_nilai_kouta = nilai_kouta

            with tab1:
                nilai_rentang_gaji["pre"] = st_tags(
                    label="Kalimat Awalan :",
                    text='Press enter to add more',
                    maxtags = 200,
                    value=data_default_nilai_rentang_gaji["pre"],
                    key='pre_gaji'
                )
                
                nilai_rentang_gaji["succ"] = st_tags(
                    label="Kalimat Akhiran :",
                    text='Press enter to add more',
                    maxtags = 200,
                    value=data_default_nilai_rentang_gaji["succ"],
                    key='succ_gaji'
                )
            with tab2:
                nilai_kouta["pre"] = st_tags(
                    label="Kalimat Awalan :",
                    text='Press enter to add more',
                    maxtags = 200,
                    value=data_default_nilai_kouta["pre"],
                    key='pre_kouta'
                )
                
                nilai_kouta["succ"] = st_tags(
                    label="Kalimat Akhiran :",
                    text='Press enter to add more',
                    maxtags = 200,
                    value=data_default_nilai_kouta["succ"],
                    key='succ_kouta'
                )
        
        # Klasifikasi Bobot Klasifikasi
        with st.container(border=True):
            st.html(f"""
                <div class="box-test">
                    <div class="line black"></div>
                    <div class="step">5</div>
                </div>
            """)

            st.write('#### Bobot Klasifikasi (%)')
            
            st.write('Bobot ini menentukan nilai apakah suatu konten Lowongan / Non-Lowongan. Pastikan total dari semua bobot tidak lebih dari 100%.')

            left, center, right = st.columns(3)

            list_bobot = ["Akun Loker"] + keywords + ["Rentang Gaji", "Kouta Lowongan"]

            # st.write(len(list_bobot))
            for index, item in enumerate(list_bobot):
                bobot_awal = 0

                if(index==0):
                    bobot_awal = 60.00
                else:
                    bobot_awal = 40/(len(list_bobot)-1)

                if config and "nilai_bobot" in config and len(list_bobot) == len(config["nilai_bobot"]):
                    bobot_awal = config["nilai_bobot"][index]

                if index % 3 == 0:
                    with left:
                        nilai_bobot.append(st.number_input(f"{item} :", value=bobot_awal, disabled=not is_superadmin))
                elif index % 3 == 1:
                    with center:
                        nilai_bobot.append(st.number_input(f"{item} :", value=bobot_awal, disabled=not is_superadmin))
                else:
                    with right:
                        nilai_bobot.append(st.number_input(f"{item} :", value=bobot_awal, disabled=not is_superadmin))
            
            total = sum(nilai_bobot)

            if round(total, 2) < 100:
                st.write(f":red[*Data Bobot Kurang dari 100%. Tambahkan nilai Bobot sebanyak {abs(round(100 - total, 2))} %.]")
            elif round(total, 2) > 100:
                st.write(f":red[*Data Bobot Lebih dari 100%. Kurangi nilai Bobot sebanyak {abs(round(100 - total, 2))} %.]")
            else:
                st.write(":green[*Total Data Bobot Sudah PAS 100%]")
        
        # Output Klasifikasi
        with st.container(border=True):
            st.html(f"""
                <div class="box-test">
                    <div class="line white"></div>
                    <div class="step">6</div>
                </div>
            """)
            
            st.subheader('Keluaran Proses (Ouput)')
    
            data_target = st.radio("Pilih Mode Output :", ["**Tampilkan Hasil**","**Masukan Database**"], horizontal=True)
            
            match data_target:
                case "**Masukan Database**":
                    db_target = st.selectbox("Pilih Tabel", list_collection + ["Create New Table"])

                    if db_target == "Create New Table":
                        collection_name = st.text_input("Masukan Nama Table :", value="hasil_proses_", placeholder="Contoh format : hasil_proses_xxx", help="Mohon Masukan Nama Database Dengan Format *hasil_proses_xxx*")
                    else:
                        collection_name = db_target
                        mode_query = st.radio("Pilih Mode", ["Insert", "Upsert"], horizontal=True)

                        if(mode_query=="Upsert"):
                            list_kolom = db[db_target].find_one()
                            
                            if list_kolom:
                                update_by = st.multiselect("Update by :", list_kolom.keys())
                            else:
                                update_by = []
                                st.error("Table Belum Memiliki Data Kolom")
                case "**Tampilkan Hasil**":
                    st.info("Data hasil proses akan langsung tampil di dashboard. Data juga dapat di download dengan format Exel.", icon=":material/info:")
                case default:
                    st.error("System Error")
    with col1:
        st.write("## Mesin Klasifikasi Text 📄 🚀")
        st.write("###### Klasifikasi Data Text tidak ter-struktur dengan akurat & presisi.")

        st.html(
            """
                <img src='https://i.ibb.co.com/mHJpjNb/mantap.gif' style='width: 100%; border-radius: 0.5rem; pointer-events: none;'/>
            """
        )
        
        with st.container(border=True):                       
            progress_bar = st.progress(0, "Progress Chunk & Split Data ...")
            sub_progress_bar = st.progress(0, "Progress Indentifikasi & Klasifikasi Text ...")
        
        left, right = st.columns(2) 

        with left:
            if st.button("Simpan Konfig", type="secondary", icon=":material/save:", use_container_width=True, disabled=not is_superadmin):
                
                collection = db["config"]
                
                data_config = {
                    "keywords": keywords,
                    "kat_akun_loker": kat_akun_loker,
                    "nilai_bobot": nilai_bobot,
                    "nilai_rentang_gaji": nilai_rentang_gaji,
                    "nilai_kouta": nilai_kouta,
                    "selected_klasifikasi_aktif": selected_klasifikasi_aktif
                }

                collection.drop()

                collection.insert_one(data_config)

                st.toast("Berhasil Menyimpan Konfigurasi!")

        with right:
            eksekusi = st.button("Jalankan Proses", type="primary", icon=":material/play_arrow:", use_container_width=True, disabled=not is_superadmin)

        if(eksekusi):
            # Proses By Data Source Type
            if mode_klasifikasi=="Proses Baru":
                match data_source:
                    case "Data Lokal" | "Export Telegram":
                        for uploaded_file in uploaded_files:
                            if(uploaded_file):
                                processed_chunks = []
                                st.toast(f"Memulai proses ... {uploaded_file.name}")

                                progress_bar.progress(30, "Membaca Kamus Data ...")
                                kd = get_kamus_data()

                                if data_source=="Data Lokal":
                                    total_rows = pd.read_excel(uploaded_file, engine='openpyxl', sheet_name="Media Sosial").shape[0]
                                else:
                                    # data_json = pd.read_json(uploaded_file, orient="records")
                                    json_data = json.load(uploaded_file)
                                    df_json = parse_json_telegram(json_data, keyword_filter_tele) 
                                    total_rows = df_json.shape[0]

                                progress_bar.progress(50, "Membaca Excel File ...")
                                progress_bar.progress(70, "Menghitung Chunk File ...")
                                num_chunks = (total_rows // chunksize) + (total_rows % chunksize > 0)

                                for i in range(num_chunks):
                                    if data_source=="Data Lokal":
                                        chunk = pd.read_excel(
                                            uploaded_file,
                                            sheet_name="Media Sosial",
                                            engine='openpyxl',
                                            skiprows=range(1, i * chunksize + 1),  # Skip rows that have already been processed
                                            nrows=chunksize,  # Read only 'chunksize' rows at a time
                                        )
                                    else:
                                        chunk = df_json.iloc[i * chunksize: (i + 1) * chunksize]

                                    progress_bar.progress((i + 1) / num_chunks, f"Memproses {i + 1}/{num_chunks} chunk data")
                                    processed_chunk = process_chunk(chunk, kd, sub_progress_bar)

                                    progress_bar.progress((i + 1) / num_chunks, f"Mengirim data ke server {i + 1}/{num_chunks} ...")

                                    match data_target:
                                        case "**Masukan Database**":
                                            if mode_query == "Insert":
                                                data_dict = processed_chunk.to_dict(orient="records")
                                                collection = db[collection_name]
                                                collection.insert_many(data_dict)
                                            else:
                                                data_dict = processed_chunk.to_dict(orient="records")
                                                collection = db[collection_name]

                                                operations = []

                                                for record in data_dict:
                                                    # Create a filter based on the update_keys
                                                    update_by = update_by if update_by else ["UUID"]
                                                    filter_query = {key: record[key] for key in update_by if key in record}
                                                    
                                                    # Ensure the record is valid for upserting
                                                    if filter_query:
                                                        operations.append(
                                                            UpdateOne(
                                                                filter_query,  # Match condition
                                                                {'$set': record},  # Update with new data
                                                                upsert=True     # Insert if not found
                                                            )
                                                        )

                                                if operations:
                                                    collection.bulk_write(operations)
                                                # st.write("mode ini blm ter-implement")
                                        case "**Tampilkan Hasil**":
                                            processed_chunks.append(processed_chunk)
                                            time.sleep(0.5)
                                    
                                    
                                if data_target=="**Tampilkan Hasil**":
                                    final_df = pd.concat(processed_chunks, ignore_index=True)
                                    preview_result(kd, final_df)
                                else:
                                    st.toast("Berhasil Memproses Klasifikasi.")
                            else:
                                st.toast("Mohon upload file excel yang ingin diproses.")
                    case "Data Server":
                        base_url = "https://api.kurasi.media/new-export/456"
                        dynamic_url = generate_url(base_url, from_date.strftime("%Y-%m-%d"), to_date.strftime("%Y-%m-%d"))
                        file_bytes = None
                        
                        progress_bar.progress(20, "Mengambil file ke server ...")

                        try:
                            response = requests.get(dynamic_url)
                            progress_bar.progress(40, "Cek Status Request ...")
                            response.raise_for_status()
                            progress_bar.progress(80, "Konversi Fle ke Format Byte ...")
                            file_bytes = BytesIO(response.content)
                            progress_bar.progress(100, "Selesai Konversi Fle ...")
                        except requests.exceptions.RequestException as e:
                            st.toast(f"Gagal mengambil data server : {e}")
                    
                        if file_bytes:
                            st.toast("Memulai proses ...")

                            progress_bar.progress(30, "Membaca Kamus Data ...")
                            kd = get_kamus_data()
                            progress_bar.progress(50, "Membaca Excel File ...")
                            total_rows = pd.read_excel(file_bytes, engine='openpyxl', sheet_name="Media Sosial").shape[0]
                            progress_bar.progress(80, "Menghitung Chunk File ...")
                            num_chunks = (total_rows // chunksize) + (total_rows % chunksize > 0)
                            
                            for i in range(num_chunks):
                                chunk = pd.read_excel(
                                    file_bytes,
                                    sheet_name="Media Sosial",
                                    engine='openpyxl',
                                    skiprows=range(1, i * chunksize + 1),  # Skip rows that have already been processed
                                    nrows=chunksize,  # Read only 'chunksize' rows at a time
                                )
                                progress_bar.progress((i + 1) / num_chunks, f"Memproses {i + 1}/{num_chunks} chunk data")
                                process_chunk(chunk, kd, sub_progress_bar)
                                
                                progress_bar.progress((i + 1) / num_chunks, f"Mengirim data ke server {i + 1}/{num_chunks} ...")
                                time.sleep(0.5)
                    case "Masukan Manual":
                        # st.info("Not Implemented")
                        # st.toast(data_text_manual)
                        progress_bar.progress(20, "Membaca Kamus Data ...")
                        kd = get_kamus_data()
                        progress_bar.progress(60, "Memproses Text ...")
                        chunk = pd.DataFrame([{"Konten": data_text_manual}])
                        result = process_chunk(chunk, kd, sub_progress_bar)

                        if data_target=="**Tampilkan Hasil**":
                            preview_result(kd, result, tipe="json")
                    case "Export Telegram":
                        for uploaded_file in uploaded_files:
                            if(uploaded_file):
                                st.toast(f"Memulai proses ... {uploaded_file.name}")

                                progress_bar.progress(30, "Membaca Kamus Data ...")
                                kd = get_kamus_data()

                                
                                json_data = json.load(uploaded_file)
                                df = parse_json_telegram(json_data, keyword_filter_tele)
                            else:
                                st.toast("Mohon upload file json yang ingin diproses.")

                    case default:
                        st.toast("This data source not implmented yet.", icon="ℹ️")

                progress_bar.progress(0, "Progress Chunk & Split Data ...")
            else:
                # Display results
                processed_chunks = []
                progress_bar.progress(10, "Membaca Kamus Data ...")
                kd = get_kamus_data()
                progress_bar.progress(35, "Mendapatkan Source Data ...")
                result_list = execute_query_builder(db_source, fields, sort_fields, sort_order_value)
                total_rows = len(result_list)
                progress_bar.progress(60, "Menghitung Chunk File ...")
                num_chunks = (total_rows // chunksize) + (total_rows % chunksize > 0)
                
                for i in range(num_chunks):
                    chunk_start = i * chunksize
                    chunk_end = chunk_start + chunksize
                    chunk = result_list[chunk_start:chunk_end]

                    chunk_df = pd.DataFrame(chunk)

                    progress_bar.progress((i + 1) / num_chunks, f"Memproses {i + 1}/{num_chunks} chunk data")
                    processed_chunk = process_chunk(chunk_df, kd, sub_progress_bar)

                    progress_bar.progress((i + 1) / num_chunks, f"Mengirim data ke server {i + 1}/{num_chunks} ...")

                    match data_target:
                        case "**Masukan Database**":
                            if mode_query == "Insert":
                                data_dict = processed_chunk.to_dict(orient="records")
                                collection = db[collection_name]
                                collection.insert_many(data_dict)
                            else:
                                data_dict = processed_chunk.to_dict(orient="records")
                                collection = db[collection_name]

                                operations = []

                                for record in data_dict:
                                    # Create a filter based on the update_keys
                                    update_by = update_by if update_by else ["UUID"]
                                    filter_query = {key: record[key] for key in update_by if key in record}
                                    
                                    # Ensure the record is valid for upserting
                                    if filter_query:
                                        operations.append(
                                            UpdateOne(
                                                filter_query,  # Match condition
                                                {'$set': record},  # Update with new data
                                                upsert=True     # Insert if not found
                                            )
                                        )

                                if operations:
                                    collection.bulk_write(operations)
                                # st.write("mode ini blm ter-implement")
                        case "**Tampilkan Hasil**":
                            processed_chunks.append(processed_chunk)
                            time.sleep(0.5)
                        
                if data_target=="**Tampilkan Hasil**":
                    final_df = pd.concat(processed_chunks, ignore_index=True)
                    preview_result(kd, final_df)
                else:
                    st.toast("Berhasil Memproses Klasifikasi.")

                progress_bar.progress(0, "Progress Chunk & Split Data ...")